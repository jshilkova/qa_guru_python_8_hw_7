import os
from zipfile import ZipFile

import openpyxl
from pypdf import PdfReader
import xlrd

import utils


def test_zip_file_names():
    with ZipFile(os.path.join(utils.TMP_PATH, "resources.zip")) as zip_file:
        resource_file_names = os.listdir(utils.RESOURCES_PATH)
        zip_file_names = zip_file.namelist()
        assert len(zip_file_names) == len(resource_file_names)
        for file in resource_file_names:
            assert file in zip_file_names


def test_zip_file_contents():
    with ZipFile(os.path.join(utils.TMP_PATH, "resources.zip")) as zip_file:
        for file in zip_file.namelist():
            original_file = os.path.join(utils.RESOURCES_PATH, file)
            assert zip_file.getinfo(file).file_size == os.path.getsize(original_file)
            with open(original_file, 'rb') as f:
                assert zip_file.read(file) == f.read()


def test_zip_file_xls():
    file_name = "file_example_XLS_10.xls"
    o_book = xlrd.open_workbook(os.path.join(utils.RESOURCES_PATH, file_name))
    with ZipFile(os.path.join(utils.TMP_PATH, "resources.zip")) as zip_file:
        zip_book = xlrd.open_workbook(file_contents=zip_file.read(file_name))
        assert zip_book.nsheets == o_book.nsheets
        assert zip_book.sheet_names() == o_book.sheet_names()
        assert zip_book.sheet_by_index(0).ncols == o_book.sheet_by_index(0).ncols
        assert zip_book.sheet_by_index(0).nrows == o_book.sheet_by_index(0).nrows
        assert zip_book.sheet_by_index(0).col_values(1) == o_book.sheet_by_index(0).col_values(1)
        assert zip_book.sheet_by_index(0).row_values(0) == o_book.sheet_by_index(0).row_values(0)
        assert zip_book.sheet_by_index(0).cell_value(1, 2) == o_book.sheet_by_index(0).cell_value(1, 2)


def test_zip_file_xlsx():
    file_name = "file_example_XLSX_50.xlsx"
    o_book = openpyxl.load_workbook(os.path.join(utils.RESOURCES_PATH, file_name))
    o_sheet = o_book.active
    mylist = list(o_sheet.iter_rows(max_row=1, values_only=True))
    print(mylist)
    with (ZipFile(os.path.join(utils.TMP_PATH, "resources.zip")) as zip_file):
        zip_book = openpyxl.load_workbook(zip_file.open(file_name))
        zip_sheet = zip_book.active
        assert zip_book.sheetnames == o_book.sheetnames
        assert zip_sheet.max_column == o_sheet.max_column
        assert zip_sheet.max_row == o_sheet.max_row
        assert [c.value for c in zip_sheet['B']] == [c.value for c in o_sheet['B']]
        assert list(zip_sheet.iter_rows(max_row=1, values_only=True)) \
               == list(o_sheet.iter_rows(max_row=1, values_only=True))
        assert zip_sheet.cell(1, 2).value == o_sheet.cell(1, 2).value


def test_zip_file_pdf():
    file_name = "Python Testing with Pytest (Brian Okken).pdf"
    o_reader = PdfReader(os.path.join(utils.RESOURCES_PATH, file_name))
    with ZipFile(os.path.join(utils.TMP_PATH, "resources.zip")) as zip_file:
        zip_reader = PdfReader(zip_file.open(file_name, 'r'))
        assert len(zip_reader.pages) == len(o_reader.pages)
        assert zip_reader.pages[10].extract_text() == o_reader.pages[10].extract_text()
        assert not zip_reader.attachments
